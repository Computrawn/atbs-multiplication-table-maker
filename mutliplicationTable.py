#! python3
# mutliplicationTable.py — An exercise in working with Excel files.

import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

N_value = int(input("Input value of N: "))


def create_table(N):
    """Creates NxN table with bolded cell data."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    font_obj = Font(bold=True)

    for row in range(1, N + 1):
        sheet[f"A{row + 1}"] = row
        sheet[f"A{row + 1}"].font = font_obj

    for column in range(1, N + 1):
        column_no = get_column_letter(column + 1)
        sheet[f"{column_no}1"] = column
        sheet[f"{column_no}1"].font = font_obj

    wb.save("multiplication_table.xlsx")


# TODO: Create loop to apply formula for cell calculations.

create_table(N_value)
