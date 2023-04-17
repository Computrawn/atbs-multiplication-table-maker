#! python3
# mutliplicationTable.py — An exercise in working with Excel files.

import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# TODO: Change N_value from user input to sysarg for shell script.
N_value = sys.argv[1]


def create_table(N_value):
    """Creates and saves an NxN multiplication table with bolded headers."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    bold = Font(bold=True)
    for row in range(1, N_value + 1):
        sheet[f"A{row + 1}"] = row
        sheet[f"A{row + 1}"].font = bold

    for column in range(1, N_value + 1):
        column_letter = get_column_letter(column + 1)
        sheet[f"{column_letter}1"] = column
        sheet[f"{column_letter}1"].font = bold

    for column in range(1, N_value + 1):
        column_letter = get_column_letter(column + 1)
        for row in range(2, N_value + 2):
            sheet[f"{column_letter}{row}"] = (
                sheet[f"{column_letter}1"].value * sheet[f"A{row}"].value
            )

    wb.save("multiplication_table.xlsx")


create_table(N_value)
