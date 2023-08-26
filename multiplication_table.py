#!/usr/bin/env python3
# mutliplicationTable.py — An exercise in working with Excel files.
# For more information, see README.md

import logging
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.DEBUG,
    filename="logging.txt",
    format="%(asctime)s -  %(levelname)s -  %(message)s",
)
logging.disable(logging.CRITICAL)  # Note out to enable logging.


def create_table(n_value):
    """Creates and saves an NxN multiplication table with bolded headers."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    bold = Font(bold=True)
    for row in range(1, n_value + 1):
        sheet[f"A{row + 1}"] = row
        sheet[f"A{row + 1}"].font = bold

    for column in range(1, n_value + 1):
        column_letter = get_column_letter(column + 1)
        sheet[f"{column_letter}1"] = column
        sheet[f"{column_letter}1"].font = bold

    for column in range(1, n_value + 1):
        column_letter = get_column_letter(column + 1)
        for row in range(2, n_value + 2):
            sheet[f"{column_letter}{row}"] = (
                sheet[f"{column_letter}1"].value * sheet[f"A{row}"].value
            )

    wb.save("multiplication_table.xlsx")


def main():
    while True:
        n_value = int(input("Please enter a positive integer: "))
        if n_value > 0:
            create_table(n_value)
            print(f"Creating {n_value} by {n_value} multiplication table.")
            return 0
        else:
            print("Not a valid entry. Try again.")


if __name__ == "__main__":
    main()
